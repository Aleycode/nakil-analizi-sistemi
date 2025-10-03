"""
Ana uygulama dosyası - Excel veri analizi ve parquet dönüştürme
Temizlenmiş versiyon
"""

import argparse
import logging
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional
import shutil

from src.core.config import (
    LOG_DOSYA,
    LOG_SEVIYE,
    ISLENMIŞ_VERI_DIZIN,
    RAPOR_DIZIN,
    PROGRAM_AYARLARI,
    OTOMATIK_ANALIZ_AYARLARI,
    EXCEL_TARIH_SUTUNLARI,
    VAKA_TIPI_ISIMLERI,
)
from src.processors.veri_isleme import VeriIsleme
from src.analyzers.nakil_analyzer import NakilAnalizcisi
import pandas as pd

# Logger yapılandırması
logging.basicConfig(
    level=getattr(logging, LOG_SEVIYE),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_DOSYA, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)


def eski_verileri_temizle() -> None:
    """
    Program başlangıcında eski verileri temizler (config ayarına göre)
    """
    try:
        if not PROGRAM_AYARLARI.get("eski_verileri_sil", True):
            logger.info("Eski veri temizleme devre dışı")
            return

        logger.info("Eski veriler temizleniyor...")

        # Processed klasörünü temizle
        if ISLENMIŞ_VERI_DIZIN.exists():
            for dosya in ISLENMIŞ_VERI_DIZIN.iterdir():
                if dosya.is_file():
                    dosya.unlink()
                    logger.debug(f"Silindi: {dosya}")
            logger.info(f"Processed klasörü temizlendi: {ISLENMIŞ_VERI_DIZIN}")

        # Reports klasörünü temizle
        if RAPOR_DIZIN.exists():
            for item in RAPOR_DIZIN.iterdir():
                if item.is_file():
                    item.unlink()
                    logger.debug(f"Dosya silindi: {item}")
                elif item.is_dir():
                    shutil.rmtree(item)
                    logger.debug(f"Klasör silindi: {item}")
            logger.info(f"Reports klasörü temizlendi: {RAPOR_DIZIN}")

        print("🧹 Eski veriler temizlendi")

    except Exception as e:
        logger.error(f"Eski veri temizleme hatası: {e}")
        print(f"⚠️  Eski veri temizleme hatası: {e}")


def gunluk_islem_yap(excel_dosya: str) -> None:
    """
    Günlük veri işleme operasyonu

    Args:
        excel_dosya: İşlenecek Excel dosyasının yolu
    """
    try:
        logger.info(f"Günlük işlem başlatılıyor: {excel_dosya}")

        # Veri işleyici oluştur
        isleyici = VeriIsleme()

        # Günlük işlemi gerçekleştir
        sonuc = isleyici.gunluk_islem(excel_dosya)

        print("✅ Günlük işlem başarıyla tamamlandı!")
        print(f"📊 İşlenen satır sayısı: {sonuc['işlenen_satir_sayisi']}")
        print(f"💾 Günlük dosya: {sonuc['gunluk_parquet']}")
        print("📋 Günlük rapor oluşturuldu")

        # Otomatik analiz kontrolü
        if OTOMATIK_ANALIZ_AYARLARI.get("gunluk_islem_sonrasi_analiz", True):
            print("\n🔄 Otomatik nakil analizi başlatılıyor...")

            # Dün analizi
            if OTOMATIK_ANALIZ_AYARLARI.get("dun_analizi", True):
                try:
                    print("📅 Dün analizi yapılıyor...")
                    gunluk_nakil_analizi_yap(gun_tipi="dun")
                    print("✅ Dün analizi tamamlandı")
                except Exception as e:
                    logger.error(f"Dün analizi hatası: {e}")
                    print(f"❌ Dün analizi hatası: {e}")

            # Bugün analizi
            if OTOMATIK_ANALIZ_AYARLARI.get("bugun_analizi", True):
                try:
                    print("\n📅 Bugün analizi yapılıyor...")
                    gunluk_nakil_analizi_yap(gun_tipi="bugun")
                    print("✅ Bugün analizi tamamlandı")
                except Exception as e:
                    logger.error(f"Bugün analizi hatası: {e}")
                    print(f"❌ Bugün analizi hatası: {e}")

            print("\n🎉 Tüm otomatik işlemler tamamlandı!")
        else:
            print("📊 Otomatik analiz devre dışı")

    except Exception as e:
        logger.error(f"Günlük işlem hatası: {e}")
        print(f"❌ Hata: {e}")
        sys.exit(1)


def gunluk_nakil_analizi_yap(
    gun_tarihi: Optional[str] = None, gun_tipi: str = "dun"
) -> None:
    """
    Günlük nakil analizi yapar

    Args:
        gun_tarihi: Analiz günü (YYYY-MM-DD formatında), None ise bugün
        gun_tipi: "dun" veya "bugun" - analiz tipini belirler
    """
    try:
        if gun_tarihi is None:
            if gun_tipi == "bugun":
                gun_tarihi = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
            else:
                gun_tarihi = datetime.now().strftime("%Y-%m-%d")
        else:
            if gun_tipi == "bugun":
                verilen_tarih = datetime.strptime(gun_tarihi, "%Y-%m-%d")
                gun_tarihi = (verilen_tarih + timedelta(days=1)).strftime("%Y-%m-%d")

        gun_datetime = datetime.strptime(gun_tarihi, "%Y-%m-%d")
        baslangic_tarihi = (gun_datetime - timedelta(days=1)).strftime("%Y-%m-%d")

        logger.info(f"Günlük nakil analizi başlatılıyor: {gun_tarihi} ({gun_tipi})")
        logger.info(f"Zaman aralığı: {baslangic_tarihi} 08:00 - {gun_tarihi} 08:00")

        analizci = NakilAnalizcisi()
        rapor = analizci.kapsamli_gunluk_analiz(gun_tarihi)
        
        if not rapor:
            print("❌ Analiz raporu oluşturulamadı, veri bulunamadı.")
            return

        print("📊 GÜNLÜK NAKİL ANALİZİ SONUÇLARI")
        print("=" * 50)
        print(f"📅 Analiz tarihi: {gun_tarihi}")
        print(f"🔄 Analiz tipi: {gun_tipi.title()}")
        print(f"⏰ Zaman aralığı: {baslangic_tarihi} 08:00 - {gun_tarihi} 08:00")
        print(f"📈 Toplam vaka sayısı: {rapor['toplam_vaka_sayisi']:,}")

        # Genel istatistikler
        if "genel_istatistikler" in rapor and rapor["genel_istatistikler"]:
            stats = rapor["genel_istatistikler"]
            print(
                f"🆕 {VAKA_TIPI_ISIMLERI['yeni_vaka_adi']} sayısı: {stats.get('yeni_vaka_sayisi', 0):,}"
            )
            print(
                f"🔄 {VAKA_TIPI_ISIMLERI['devreden_vaka_adi']} sayısı: {stats.get('devreden_vaka_sayisi', 0):,}"
            )
            print(
                f"📊 {VAKA_TIPI_ISIMLERI['yeni_vaka_adi']} oranı: {stats.get('yeni_vaka_yuzde', 0):.1f}%"
            )
            print(
                f"📊 {VAKA_TIPI_ISIMLERI['devreden_vaka_adi']} oranı: {stats.get('devreden_vaka_yuzde', 0):.1f}%"
            )

        # İl grupları özetini göster
        if "il_gruplari" in rapor:
            for il_grup, il_veri in rapor["il_gruplari"].items():
                if il_veri:
                    # Daha anlaşılır isimler göster
                    if il_grup == "Butun_Bolgeler":
                        print(f"\n📍 Bütün Bölgeler analiz edildi")
                    elif il_grup == "Sevk_Vakalar":
                        print(f"\n📍 Sevk Vakaları analiz edildi")
                    elif il_grup == "Yerel_Vakalar":
                        print(f"\n📍 Yerel Vakalar analiz edildi")
                    else:
                        print(f"\n📍 {il_grup} analiz edildi")

        print(
            f"\n💾 Detaylı rapor: {RAPOR_DIZIN}/kapsamli_gunluk_analiz_{gun_tarihi}.json"
        )
        print("📊 Grafikler reports klasöründe oluşturuldu")

        # PDF raporu bilgisi
        if "pdf_raporu" in rapor:
            print(f"📄 PDF raporu oluşturuldu: {rapor['pdf_raporu']}")

        # Excel raporu oluştur
        excel_raporu_olustur(rapor, gun_tarihi)

        # Tüm grafiklerin tek PDF sayfasında birleştirilmesi
        try:
            from src.generators.grafik_olusturucu import GrafikOlusturucu
            go = GrafikOlusturucu()
            pdf_path = go.tum_grafikleri_pdfde_birlestir(gun_tarihi)
            if pdf_path:
                print(f"📄 Tüm grafikler tek PDF sayfasında: {pdf_path}")
        except Exception as e:
            logger.warning(f"Grafikleri PDF'de birleştirme hatası: {e}")

    except Exception as e:
        logger.error(f"Günlük nakil analizi hatası: {e}")
        print(f"❌ Hata: {e}")


def tarih_formati_uygula(workbook, sheet_name):
    """Excel sayfasındaki tarih sütunlarına dd-mm-yyyy hh:mm formatı uygular"""
    try:
        from openpyxl.styles import NamedStyle

        # Tarih formatı stili oluştur
        tarih_stili = NamedStyle(name="tarih_stili")
        tarih_stili.number_format = "DD-MM-YYYY HH:MM"

        # Workbook'a stil ekle
        if "tarih_stili" not in workbook.named_styles:
            workbook.add_named_style(tarih_stili)

        # Sayfadaki tarih sütunlarını formatla
        worksheet = workbook[sheet_name]
        tarih_sutunlari = ["basvuru_tarihi", "vaka_tarihi", "nakil_tarihi"]

        # İlk satırdan sütun başlıklarını al
        basliklar = []
        for cell in worksheet[1]:
            if cell.value:
                basliklar.append(cell.value.lower())

        # Tarih sütunlarının indekslerini bul
        tarih_indeksleri = []
        for i, baslik in enumerate(basliklar):
            tarih_sut_var = any(
                tarih_sut in str(baslik).lower() for tarih_sut in tarih_sutunlari
            )
            if tarih_sut_var:
                tarih_indeksleri.append(i + 1)  # Excel 1-indexed

        # Tarih formatını uygula
        for col_idx in tarih_indeksleri:
            for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell.style = "tarih_stili"

    except Exception as e:
        logging.warning(f"Tarih formatı uygulanırken hata: {e}")


def excel_raporu_olustur(rapor: dict, gun_tarihi: str) -> None:
    """
    Analiz verilerini Excel formatında reports klasörüne kaydeder

    Args:
        rapor: Analiz raporu verileri
        gun_tarihi: Analiz günü
    """

    def tarih_formati_uygula(workbook, worksheet_name: str):
        """Excel sayfasındaki tarih sütunlarına dd-mm-yyyy hh:mm formatı uygular"""
        from openpyxl.styles import NamedStyle

        # Tarih formatı tanımla
        tarih_stili = NamedStyle(name="tarih_formati")
        tarih_stili.number_format = "DD-MM-YYYY HH:MM"

        # Eğer stil henüz eklenmemişse workbook'a ekle
        if "tarih_formati" not in workbook.named_styles:
            workbook.add_named_style(tarih_stili)

        worksheet = workbook[worksheet_name]

        # İlk satırdaki başlık sütunlarını bul
        if worksheet.max_row > 0:
            header_row = list(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )[0]

            # Tarih sütunlarının indekslerini bul
            tarih_sutun_indeksleri = []
            for i, header in enumerate(header_row):
                if header in EXCEL_TARIH_SUTUNLARI:
                    tarih_sutun_indeksleri.append(i + 1)  # Excel 1-indexed

            # Tarih sütunlarına format uygula
            for col_idx in tarih_sutun_indeksleri:
                for row in range(2, worksheet.max_row + 1):  # Başlık hariç
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value and cell.value != "":
                        cell.style = "tarih_formati"

    try:
        # Ana veriyi oku
        analizci = NakilAnalizcisi()
        df_tum_veri = analizci.veriyi_oku()  # Tüm ham veri

        # Günlük zaman aralığında filtrele (analiz için)
        df_gunluk = analizci.gunluk_zaman_araligi_filtrele(df_tum_veri, gun_tarihi)
        df_gunluk = analizci.vaka_tipi_belirle(df_gunluk, gun_tarihi)

        # Tarih klasörü oluştur ve Excel dosyası oluştur
        tarih_klasor = RAPOR_DIZIN / gun_tarihi
        tarih_klasor.mkdir(parents=True, exist_ok=True)
        excel_dosya = tarih_klasor / f"nakil_analiz_raporu_{gun_tarihi}.xlsx"

        with pd.ExcelWriter(excel_dosya, engine="openpyxl") as writer:
            # Ham veri sayfası (TÜM VERİ)
            df_tum_veri.to_excel(writer, sheet_name="Ham_Veri", index=False)

            # Yeni vakalar sayfası
            yeni_vakalar = df_gunluk[df_gunluk["vaka_tipi"] == "Yeni Vaka"].copy()
            if len(yeni_vakalar) > 0:
                yeni_vakalar.to_excel(writer, sheet_name="Yeni_Vakalar", index=False)

            # Devreden vakalar sayfası
            devreden_vakalar = df_gunluk[
                df_gunluk["vaka_tipi"] == "Devreden Vaka"
            ].copy()
            if len(devreden_vakalar) > 0:
                devreden_vakalar.to_excel(
                    writer, sheet_name="Devreden_Vakalar", index=False
                )

            # Filtrelenmiş vakalar (klinik analizine dahil edilen)
            analizci_temp = NakilAnalizcisi()
            filtrelenmis_vakalar = analizci_temp.klinik_filtrele(df_gunluk)
            if len(filtrelenmis_vakalar) > 0:
                filtrelenmis_vakalar.to_excel(
                    writer, sheet_name="Filtrelenmis_Vakalar", index=False
                )

            # İl grupları için doğru veri kullan (sadece geçerli vakalar)
            il_gruplari = analizci.il_bazinda_grupla(df_gunluk)

            # İl içi vakalar (sadece geçerli vakalar - filtrelenmiş hariç)
            il_ici_gecerli = il_gruplari["Il_Ici"][
                il_gruplari["Il_Ici"]["vaka_tipi"].isin(["Yeni Vaka", "Devreden Vaka"])
            ]
            if len(il_ici_gecerli) > 0:
                il_ici_gecerli.to_excel(
                    writer, sheet_name="Il_Ici_Vakalar", index=False
                )

            # İl dışı vakalar (sadece geçerli vakalar - filtrelenmiş hariç)
            il_disi_gecerli = il_gruplari["Il_Disi"][
                il_gruplari["Il_Disi"]["vaka_tipi"].isin(["Yeni Vaka", "Devreden Vaka"])
            ]
            if len(il_disi_gecerli) > 0:
                il_disi_gecerli.to_excel(
                    writer, sheet_name="Il_Disi_Vakalar", index=False
                )

            # Bütün bölgeler
            if len(il_gruplari["Butun_Bolgeler"]) > 0:
                il_gruplari["Butun_Bolgeler"].to_excel(
                    writer, sheet_name="Butun_Bolgeler", index=False
                )

            # Özet istatistikler
            ozet_data = []
            if "genel_istatistikler" in rapor and rapor["genel_istatistikler"]:
                stats = rapor["genel_istatistikler"]
                ozet_data.extend(
                    [
                        ["Metric", "Value"],
                        ["Toplam Vaka Sayısı", rapor["toplam_vaka_sayisi"]],
                        ["Yeni Vaka Sayısı", stats.get("yeni_vaka_sayisi", 0)],
                        ["Devreden Vaka Sayısı", stats.get("devreden_vaka_sayisi", 0)],
                        ["Yeni Vaka Oranı (%)", stats.get("yeni_vaka_yuzde", 0)],
                        [
                            "Devreden Vaka Oranı (%)",
                            stats.get("devreden_vaka_yuzde", 0),
                        ],
                        ["Analiz Tarihi", gun_tarihi],
                        ["Analiz Zamanı", rapor.get("analiz_zamani", "")],
                    ]
                )

                ozet_df = pd.DataFrame(ozet_data[1:], columns=ozet_data[0])
                ozet_df.to_excel(writer, sheet_name="Ozet_Istatistikler", index=False)

            # Tarih formatını tüm sayfalara uygula
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                tarih_formati_uygula(workbook, sheet_name)

        print(f"📊 Excel raporu oluşturuldu: {excel_dosya}")
        logger.info(f"Excel raporu kaydedildi: {excel_dosya}")

    except Exception as e:
        logger.error(f"Excel raporu oluşturma hatası: {e}")
        print(f"❌ Excel raporu oluşturma hatası: {e}")


def parquet_excel_donustur():
    """Parquet dosyalarını Excel formatına dönüştürür"""

    def tarih_formati_uygula(workbook, worksheet_name: str):
        """Excel sayfasındaki tarih sütunlarına dd-mm-yyyy hh:mm formatı uygular"""
        from openpyxl.styles import NamedStyle

        # Tarih formatı tanımla
        tarih_stili = NamedStyle(name="tarih_formati")
        tarih_stili.number_format = "DD-MM-YYYY HH:MM"

        # Eğer stil henüz eklenmemişse workbook'a ekle
        if "tarih_formati" not in workbook.named_styles:
            workbook.add_named_style(tarih_stili)

        worksheet = workbook[worksheet_name]

        # İlk satırdaki başlık sütunlarını bul
        if worksheet.max_row > 0:
            header_row = list(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )[0]

            # Tarih sütunlarının indekslerini bul
            tarih_sutun_indeksleri = []
            for i, header in enumerate(header_row):
                if header in EXCEL_TARIH_SUTUNLARI:
                    tarih_sutun_indeksleri.append(i + 1)  # Excel 1-indexed

            # Tarih sütunlarına format uygula
            for col_idx in tarih_sutun_indeksleri:
                for row in range(2, worksheet.max_row + 1):  # Başlık hariç
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value and cell.value != "":
                        cell.style = "tarih_formati"

    try:
        # İşlenmiş veri klasöründeki parquet dosyalarını listele
        parquet_dosyalar = list(ISLENMIŞ_VERI_DIZIN.glob("*.parquet"))

        if not parquet_dosyalar:
            print("❌ İşlenmiş veri klasöründe parquet dosyası bulunamadı!")
            return

        print("📁 Mevcut parquet dosyaları:")
        for i, dosya in enumerate(parquet_dosyalar, 1):
            dosya_boyut = dosya.stat().st_size / (1024 * 1024)  # MB
            print(f"{i}. {dosya.name} ({dosya_boyut:.2f} MB)")

        print(f"{len(parquet_dosyalar)+1}. Tümünü dönüştür")
        print(f"{len(parquet_dosyalar)+2}. Ana menüye dön")

        try:
            secim = int(input(f"Seçim (1-{len(parquet_dosyalar)+2}): ").strip())

            if 1 <= secim <= len(parquet_dosyalar):
                # Tek dosya dönüştür
                secilen_dosya = parquet_dosyalar[secim - 1]
                _tek_parquet_donustur(secilen_dosya)

            elif secim == len(parquet_dosyalar) + 1:
                # Tümünü dönüştür
                for dosya in parquet_dosyalar:
                    _tek_parquet_donustur(dosya)

            elif secim == len(parquet_dosyalar) + 2:
                return
            else:
                print("❌ Geçersiz seçim!")

        except ValueError:
            print("❌ Lütfen geçerli bir sayı girin!")

    except Exception as e:
        logger.error(f"Parquet Excel dönüştürme hatası: {e}")
        print(f"❌ Hata: {e}")


def _tek_parquet_donustur(parquet_dosya: Path):
    """Tek bir parquet dosyasını Excel'e dönüştürür"""
    try:
        # Parquet dosyasını oku
        df = pd.read_parquet(parquet_dosya)

        # Excel dosya adını oluştur
        excel_dosya = RAPOR_DIZIN / f"{parquet_dosya.stem}.xlsx"

        # Excel'e kaydet
        with pd.ExcelWriter(excel_dosya, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Data", index=False)

            # Özet sayfa ekle
            ozet_data = [
                ["Metric", "Value"],
                ["Toplam Satır", len(df)],
                ["Toplam Sütun", len(df.columns)],
                [
                    "Dosya Boyutu (MB)",
                    f"{parquet_dosya.stat().st_size / (1024*1024):.2f}",
                ],
                ["Dönüştürme Tarihi", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ]

            ozet_df = pd.DataFrame(ozet_data[1:], columns=ozet_data[0])
            ozet_df.to_excel(writer, sheet_name="Summary", index=False)

            # Tarih formatını tüm sayfalara uygula
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                tarih_formati_uygula(workbook, sheet_name)

        print(f"✅ Başarılı: {excel_dosya.name}")
        print(f"📊 Veri okundu: {len(df)} satır, {len(df.columns)} sütun")
        print(
            f"💾 Excel dosya boyutu: {excel_dosya.stat().st_size / (1024*1024):.2f} MB"
        )

    except Exception as e:
        logger.error(f"Tek parquet dönüştürme hatası: {e}")
        print(f"❌ {parquet_dosya.name} dönüştürülemedi: {e}")


def menu_gunluk_islem():
    """Günlük veri işleme menüsü"""
    print("\n📥 GÜNLÜK VERİ İŞLEME")
    print("-" * 40)

    # Raw klasöründeki Excel dosyalarını listele
    raw_klasor = Path("data/raw")
    excel_dosyalar = []

    if raw_klasor.exists():
        # Excel dosyalarını bul
        xls_dosyalar = list(raw_klasor.glob("*.xls"))
        xlsx_dosyalar = list(raw_klasor.glob("*.xlsx"))
        excel_dosyalar = xls_dosyalar + xlsx_dosyalar

    if not excel_dosyalar:
        print("❌ data/raw klasöründe Excel dosyası bulunamadı!")
        print("💡 Lütfen Excel dosyalarınızı data/raw klasörüne koyun.")
        return

    print("📁 Raw klasöründeki Excel dosyaları:")
    for i, dosya in enumerate(excel_dosyalar, 1):
        dosya_boyut = dosya.stat().st_size / (1024 * 1024)  # MB
        print(f"{i}. {dosya.name} ({dosya_boyut:.2f} MB)")

    print(f"{len(excel_dosyalar)+1}. Yeni dosya yolu gir")
    print(f"{len(excel_dosyalar)+2}. Ana menüye dön")

    try:
        secim = int(input(f"Seçim (1-{len(excel_dosyalar)+2}): ").strip())

        if 1 <= secim <= len(excel_dosyalar):
            secilen_dosya = excel_dosyalar[secim - 1]
            print(f"📁 Seçilen dosya: {secilen_dosya.name}")
            gunluk_islem_yap(str(secilen_dosya))

        elif secim == len(excel_dosyalar) + 1:
            dosya_yolu = input("Excel dosya yolu: ").strip()
            if dosya_yolu:
                gunluk_islem_yap(dosya_yolu)
            else:
                print("❌ Dosya yolu boş olamaz!")

        elif secim == len(excel_dosyalar) + 2:
            return
        else:
            print("❌ Geçersiz seçim!")

    except ValueError:
        print("❌ Lütfen geçerli bir sayı girin!")


def menu_gunluk_nakil_analizi():
    """Günlük nakil analizi menüsü - Güncellenmiş"""
    print("\n📅 GÜNLÜK NAKİL ANALİZİ")
    print("-" * 40)
    print("1. Dün için analiz (Normal: Dün 08:00 - Bugün 08:00)")
    print("2. Bugün için analiz (Bugün 08:00 - Yarın 08:00)")
    print("3. Belirli gün için analiz")

    secim = input("Seçim (1-3): ").strip()

    if secim == "1":
        gunluk_nakil_analizi_yap(gun_tipi="dun")

    elif secim == "2":
        gunluk_nakil_analizi_yap(gun_tipi="bugun")

    elif secim == "3":
        tarih = input("Analiz tarihi (YYYY-MM-DD): ").strip()
        if tarih:
            gun_tipi = input("Analiz tipi (dun/bugun): ").strip().lower()
            if gun_tipi in ["dun", "bugun"]:
                gunluk_nakil_analizi_yap(tarih, gun_tipi)
            else:
                print("❌ Geçersiz analiz tipi! 'dun' veya 'bugun' yazın.")
        else:
            print("❌ Tarih gerekli!")
    else:
        print("❌ Geçersiz seçim!")


def menu_yardim():
    """Yardım menüsü"""
    print("\n❓ YARDIM")
    print("-" * 40)
    print("📥 Günlük Veri İşleme:")
    print("   • Excel dosyalarını parquet formatına dönüştürür")
    print("   • Duplikasyon kontrolü yapar")
    print("   • Veri temizleme işlemlerini gerçekleştirir")
    print()
    print("📊 Günlük Nakil Analizi:")
    print("   • Dün: Normal günlük analiz (Dün 08:00 - Bugün 08:00)")
    print("   • Bugün: Genişletilmiş analiz (Bugün 08:00 - Yarın 08:00)")
    print("   • Yeni vaka / Devreden vaka sınıflandırması")
    print("   • İl içi/dışı gruplandırması")
    print("   • Otomatik Excel raporu oluşturma")
    print()
    print("🔄 Parquet Excel Dönüştürme:")
    print("   • Parquet dosyalarını Excel formatında inceleyin")
    print("   • Ham veri + özet sayfalar")
    print()
    print("📁 Dosya Yapısı:")
    print("   • data/raw/ : Giriş Excel dosyaları")
    print("   • data/processed/ : İşlenmiş parquet dosyaları")
    print("   • reports/ : Raporlar ve grafikler")


def console_menu():
    """İnteraktif console menü"""
    # Program başlarken eski verileri temizle
    eski_verileri_temizle()

    while True:
        print("\n" + "=" * 50)
        print("🏥 NAKİL VERİ ANALİZ SİSTEMİ")
        print("=" * 50)
        print("1. 📥 Günlük veri işleme (Excel → Parquet)")
        print("2. 📊 Günlük nakil analizi")
        print("3. 🔄 Parquet → Excel dönüştürme")
        print("4. ❓ Yardım")
        print("5. 🚪 Çıkış")

        try:
            secim = input("\nSeçiminizi yapın (1-5): ").strip()

            if secim == "1":
                menu_gunluk_islem()
            elif secim == "2":
                menu_gunluk_nakil_analizi()
            elif secim == "3":
                parquet_excel_donustur()
            elif secim == "4":
                menu_yardim()
            elif secim == "5":
                print("👋 Görüşmek üzere!")
                break
            else:
                print("❌ Geçersiz seçim! Lütfen 1-5 arası bir sayı girin.")

        except KeyboardInterrupt:
            print("\n\n👋 Çıkış yapılıyor...")
            break
        except Exception as e:
            logger.error(f"Menü hatası: {e}")
            print(f"❌ Beklenmeyen hata: {e}")


def main():
    """Ana fonksiyon"""
    parser = argparse.ArgumentParser(
        description="Excel veri analizi ve parquet dönüştürme"
    )

    parser.add_argument(
        "--gunluk-islem", type=str, help="Günlük veri işleme için Excel dosya yolu"
    )
    parser.add_argument(
        "--analiz", type=str, help="Günlük nakil analizi için tarih (YYYY-MM-DD)"
    )
    parser.add_argument(
        "--gun-tipi",
        type=str,
        choices=["dun", "bugun"],
        default="dun",
        help="Analiz tipi",
    )

    args = parser.parse_args()

    try:
        if args.gunluk_islem:
            gunluk_islem_yap(args.gunluk_islem)
        elif args.analiz:
            gunluk_nakil_analizi_yap(args.analiz, args.gun_tipi)
        else:
            # Parametre olmadan çalıştırıldıysa console menüyü başlat
            console_menu()

    except Exception as e:
        logger.error(f"Ana program hatası: {e}")
        print(f"❌ Program hatası: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
